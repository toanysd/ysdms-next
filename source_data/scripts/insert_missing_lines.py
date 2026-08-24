#!/usr/bin/env python3
# =============================================================
# insert_missing_lines.py — Insert order_lines that were skipped
#   in R6-S2 due to NOT_FOUND product codes (now resolved by stubs)
# Phase R7-S2 — ysdms-nextgen
# [local-only, not for deployment]
# PE Directive #49
# Input : source_data/parse_output_dryrun_v2.json
# Output: source_data/missing_lines_result.json
# =============================================================

import os, re, sys, json, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import requests

# ── CONFIG ────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_DIR = os.path.join(SCRIPT_DIR, "..")
INPUT_JSON = os.path.join(SOURCE_DIR, "parse_output_dryrun_v2.json")
OUTPUT_JSON = os.path.join(SOURCE_DIR, "missing_lines_result.json")

SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://iirezrszalmecsslbruo.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
REST_URL = f"{SUPABASE_URL}/rest/v1"
HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Accept": "application/json",
    "Prefer": "return=representation",
}


# ── HELPERS ───────────────────────────────────────────────────

def rest_get(table: str, params: dict) -> list:
    r = requests.get(f"{REST_URL}/{table}", headers=HEADERS, params=params, timeout=15)
    r.raise_for_status()
    return r.json()


def rest_insert(table: str, data: dict) -> dict:
    r = requests.post(f"{REST_URL}/{table}", headers=HEADERS, json=data, timeout=15)
    if r.status_code not in (200, 201):
        raise Exception(f"INSERT {table} failed ({r.status_code}): {r.text}")
    result = r.json()
    return result[0] if isinstance(result, list) else result


def fetch_product_map() -> dict:
    """product_code (upper) → product_id."""
    products = []
    offset = 0
    while True:
        batch = rest_get("products", {"select": "product_id,product_code", "limit": 1000, "offset": offset})
        products.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    return {(p["product_code"] or "").strip().upper(): p["product_id"] for p in products if p.get("product_code")}


def fetch_order_map() -> dict:
    """order_no → order_id for imported orders."""
    orders = []
    offset = 0
    while True:
        batch = rest_get("orders", {
            "select": "order_id,order_no",
            "notes": "like.%YSDトレー受注一覧%",
            "limit": 1000,
            "offset": offset,
        })
        orders.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    return {o["order_no"]: o["order_id"] for o in orders if o.get("order_no")}


def normalize_code(raw: str) -> str:
    return raw.replace("\uff0d", "-").replace("\u2010", "-").strip()


# ── MAIN ──────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", default=False)
    args = parser.parse_args()

    print("=" * 60)
    print(f"R7-S2 — Insert Missing Lines ({'DRY-RUN' if args.dry_run else 'LIVE INSERT'})")
    print(f"Input : {os.path.abspath(INPUT_JSON)}")
    print(f"Output: {os.path.abspath(OUTPUT_JSON)}")
    print("=" * 60)

    if not SUPABASE_KEY:
        sys.exit("ERROR: SUPABASE_SERVICE_KEY not set")

    # Load dryrun v2 data
    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)

    orders_data = data["orders"]

    # Collect all NOT_FOUND lines across all orders
    # We need to find lines that were NOT inserted in R6-S2
    # These are lines where product_code was NOT_FOUND at import time
    # We re-match them now against the updated products table
    
    # But wait — the dryrun_v2.json only has FOUND lines in orders[].lines
    # NOT_FOUND lines were never added to orders[].lines
    # We need to re-parse from Excel OR use a different approach
    
    # Better approach: re-read the v1 dryrun which has ALL lines
    # OR: re-run parse with updated product map
    
    # Simplest: re-run the parse logic from parse_orders_v2.py 
    # but only for NOT_FOUND lines, using updated product map
    
    print("\n  Loading parse_not_found.json for code list...")
    nf_path = os.path.join(SOURCE_DIR, "parse_not_found.json")
    with open(nf_path, "r", encoding="utf-8") as f:
        not_found_codes = json.load(f)
    nf_code_set = {normalize_code(r["product_code_raw"]).upper() for r in not_found_codes}
    print(f"    → {len(nf_code_set)} NOT_FOUND codes to re-check")

    print("  Fetching updated product map (with stubs)...")
    product_map = fetch_product_map()
    print(f"    → {len(product_map)} products (including new stubs)")

    print("  Fetching order map (imported orders)...")
    order_map = fetch_order_map()
    print(f"    → {len(order_map)} imported orders")

    # Re-parse Excel to find NOT_FOUND lines with their order context
    print("  Re-parsing Excel for NOT_FOUND lines...")
    import openpyxl
    
    EXCEL_PATH = os.path.join(SOURCE_DIR, "YSDトレー受注一覧（改2）4-22.xlsx")
    PART_NO_COLS = [1, 4, 7, 10, 13, 16, 19, 22, 25]
    QTY_COLS     = [2, 5, 8, 11, 14, 17, 20, 23, 26]
    YEAR = 2026

    # Import helpers from parse_orders_v2 logic
    def parse_date(val):
        if val is None: return None
        m = re.search(r'(\d{1,2})/(\d{1,2})', str(val).strip())
        if m:
            month, day = int(m.group(1)), int(m.group(2))
            if 1 <= month <= 12 and 1 <= day <= 31:
                return f"{YEAR}-{month:02d}-{day:02d}"
        return None

    def parse_qty_notes(val):
        if val is None: return None, None
        val_str = str(val).strip()
        if val_str in ("", "nan", "None"): return None, None
        m = re.match(r'^([\d,\.]+)\s*(.*)', val_str)
        if m:
            try: qty = int(float(m.group(1).replace(',', '')))
            except: qty = None
            notes = m.group(2).strip() or None
            return qty, notes
        return None, val_str

    def is_stray_note(text):
        if not text: return True
        text = str(text).strip()
        if not text or text in ("nan", "None"): return True
        if re.search(r'(より|前倒|後倒|出荷|引取|午前|午後|至急|検査|休み|確認|変更|祝日|休業)', text): return True
        if re.fullmatch(r'[^\x00-\x7F]+', text): return True
        if re.fullmatch(r'\d{1,2}/\d{1,2}.*', text) and not re.match(r'^[A-Z0-9]', text): return True
        return False

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Track lines per (date, company) for line_no assignment
    # We need to know the TOTAL line count (FOUND + NOT_FOUND) per order
    # to assign correct line_no that doesn't conflict with existing lines
    
    missing_lines = []
    current_date = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if not row or len(row) < 3:
            continue
        date_val = row[0]
        if date_val is not None:
            parsed = parse_date(str(date_val))
            if parsed:
                current_date = parsed
        if not current_date:
            continue

        for pn_col, qty_col in zip(PART_NO_COLS, QTY_COLS):
            if pn_col >= len(row):
                continue
            raw_code = row[pn_col]
            if raw_code is None:
                continue
            code_str = str(raw_code).strip()
            if not code_str or code_str == "nan":
                continue
            if is_stray_note(code_str):
                continue

            normalized = normalize_code(code_str)
            
            # Only process codes that were NOT_FOUND in R6-S2
            if normalized.upper() not in nf_code_set:
                continue

            # Try to match now (after stubs created)
            product_id = product_map.get(normalized.upper())
            if not product_id:
                continue  # still not found even after stubs

            # Get company for this product to build order_no
            # We need company_code — get from product_map's full data
            # Actually we need to look up order_no which requires company_code
            # Let's get company from the product
            qty_raw = row[qty_col] if qty_col < len(row) else None
            qty, notes = parse_qty_notes(qty_raw)

            if qty is None or qty <= 0:
                continue

            missing_lines.append({
                "order_date": current_date,
                "product_code": normalized,
                "product_id": product_id,
                "quantity": qty,
                "notes": notes,
                "source_row": row_idx,
            })

    wb.close()
    print(f"    → {len(missing_lines)} recoverable lines found")

    # Now we need to find the order_id for each line
    # order_no = ORD-{date_no_dash}-{company_code}
    # But we need company_code from product → company lookup
    print("  Fetching product→company mapping...")
    prod_company = {}
    offset = 0
    while True:
        batch = rest_get("products", {"select": "product_id,company_id", "limit": 1000, "offset": offset})
        for p in batch:
            prod_company[p["product_id"]] = p.get("company_id")
        if len(batch) < 1000:
            break
        offset += 1000

    company_code_map = {}
    companies = rest_get("companies", {"select": "company_id,company_code", "limit": 5000})
    for c in companies:
        company_code_map[c["company_id"]] = c.get("company_code", "")

    # Match each line to its order
    stats = {
        "total_recoverable": len(missing_lines),
        "recovered": 0,
        "order_not_found": 0,
        "line_conflict": 0,
        "errors": 0,
        "dry_run": args.dry_run,
    }

    recovered_details = []

    for line in missing_lines:
        company_id = prod_company.get(line["product_id"])
        company_code = company_code_map.get(company_id, "UNK") if company_id else "UNK"
        order_no = f"ORD-{line['order_date'].replace('-', '')}-{company_code}"
        order_id = order_map.get(order_no)

        if not order_id:
            stats["order_not_found"] += 1
            continue

        # Get next available line_no for this order
        existing_lines = rest_get("order_lines", {
            "select": "line_no",
            "order_id": f"eq.{order_id}",
            "order": "line_no.desc",
            "limit": 1,
        })
        next_line_no = (existing_lines[0]["line_no"] + 1) if existing_lines else 1

        if args.dry_run:
            stats["recovered"] += 1
            recovered_details.append({
                "order_no": order_no,
                "product_code": line["product_code"],
                "quantity": line["quantity"],
                "line_no": next_line_no,
                "action": "INSERT (dry-run)",
            })
        else:
            try:
                row_data = {
                    "order_id": order_id,
                    "product_id": line["product_id"],
                    "line_no": next_line_no,
                    "quantity": line["quantity"],
                    "unit": "PCS",
                    "line_status": "NEW",
                    "notes": f"R7-S2 recovery | raw: {line['product_code']}" +
                             (f" | {line['notes']}" if line.get("notes") else ""),
                }
                rest_insert("order_lines", row_data)
                stats["recovered"] += 1
                recovered_details.append({
                    "order_no": order_no,
                    "product_code": line["product_code"],
                    "quantity": line["quantity"],
                    "action": "INSERTED",
                })
            except Exception as e:
                stats["errors"] += 1
                recovered_details.append({
                    "order_no": order_no,
                    "product_code": line["product_code"],
                    "action": "ERROR",
                    "reason": str(e)[:200],
                })

    still_not_found = stats["total_recoverable"] - stats["recovered"] - stats["order_not_found"] - stats["errors"]

    output = {
        **stats,
        "still_not_found": still_not_found,
        "details_sample": recovered_details[:20],
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*60}")
    print(f"  RESULTS ({'DRY-RUN' if args.dry_run else 'LIVE'})")
    print(f"{'='*60}")
    for k, v in stats.items():
        if k != "dry_run":
            print(f"  {k:<30s}: {v}")
    print(f"  still_not_found             : {still_not_found}")
    print(f"{'='*60}")

    if recovered_details:
        print(f"\n  Sample recovered lines (first 10):")
        for r in recovered_details[:10]:
            print(f"    {r['order_no']:<30s} {r['product_code']:<20s} qty={r.get('quantity','-')}")

    print(f"\n  Output: {os.path.abspath(OUTPUT_JSON)}")


if __name__ == "__main__":
    main()
