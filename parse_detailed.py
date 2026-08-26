import xlrd, os, glob, sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# The 7 target forms:
targets = [
    "F 日報兼不適合製品記録書（成形最新版）2-20-4.xls",
    "F TE用日報兼不適合製品記録書（成形）1-157-8.xls",  # Matching what we saw
    "F 一般注文書兼指示書.xls",
    "F 成形機工程表.xls",
    "3号機成形条件一覧表.xls",
    "F 不適合(製品)是正報告書(客先提出用）.xls",
    "F 運転日報.xls"
]

dirs = [
    r'D:\AntiGravity_Workspace\apps\ysdms-nextgen\source_data\Form lien quan',
    r'D:\AntiGravity_Workspace\apps\ysdms-nextgen\source_data\ISO(2026見直し済み）\文書類　共通\フォーマット'
]

found_files = []
for d in dirs:
    for root, _, files in os.walk(d):
        for f in files:
            if any(t in f for t in targets):
                found_files.append(os.path.join(root, f))

# Let's make sure we only pick one of each target if multiple found
final_files = []
for t in targets:
    for f in found_files:
        if t in f:
            final_files.append(f)
            break

def parse_xls_detailed(filepath):
    try:
        print(f"\n{'='*80}")
        print(f"FILE: {os.path.basename(filepath)}")
        if filepath.endswith('.xls'):
            wb = xlrd.open_workbook(filepath, encoding_override='shift-jis')
            print(f"Sheets: {wb.sheet_names()}")
            for name in wb.sheet_names()[:3]:
                ws = wb.sheet_by_name(name)
                print(f"\n  Sheet: '{name}' ({ws.nrows} total rows x {ws.ncols} total cols)")
                for r in range(min(30, ws.nrows)):
                    vals = []
                    for c in range(ws.ncols):
                        v = ws.cell_value(r, c)
                        if v and str(v).strip():
                            val_str = str(v).replace('\n', ' ').strip()
                            vals.append(f"[{c}]{val_str}")
                    if vals:
                        print(f"    Row {r+1}: {' | '.join(vals)}")
        elif filepath.endswith('.xlsx'):
            wb = openpyxl.load_workbook(filepath, data_only=True)
            print(f"Sheets: {wb.sheetnames}")
            for name in wb.sheetnames[:3]:
                ws = wb[name]
                print(f"\n  Sheet: '{name}' ({ws.max_row} total rows x {ws.max_column} total cols)")
                for r in range(1, min(31, ws.max_row + 1)):
                    vals = []
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(row=r, column=c).value
                        if v and str(v).strip():
                            val_str = str(v).replace('\n', ' ').strip()
                            vals.append(f"[{c-1}]{val_str}")
                    if vals:
                        print(f"    Row {r}: {' | '.join(vals)}")
    except Exception as e:
        print(f"ERROR {filepath}: {e}")

for f in final_files:
    parse_xls_detailed(f)
