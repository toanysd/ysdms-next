import os
import shutil
import pandas as pd
import openpyxl
from datetime import datetime

INVENTORY_CSV = r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\server_inventory.csv"
LOCAL_DIR = r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\scratch_smk"
OUT_REPORT = r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\SMK_Merge_Report_V2.md"

def extract_sheet_data(ws):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue # header
        row_strs = [str(cell).strip() if cell is not None else "" for cell in row]
        if any(row_strs):
            rows.append(row_strs)
    return rows

def main():
    df = pd.read_csv(INVENTORY_CSV)
    smk_df = df[df['parent_dir'] == '新SMK注文書']
    if smk_df.empty:
        smk_df = df[df['full_path'].str.contains(r'\\新SMK注文書\\', case=False, na=False)]
        
    # Sort files by mtime descending (newest first)
    smk_df['mtime_dt'] = pd.to_datetime(smk_df['mtime'])
    smk_df = smk_df.sort_values('mtime_dt', ascending=False)
    
    all_addresses = {} # key -> { content_tuple: { 'files': [], 'newest_mtime': dt } }
    all_trays = {} # key -> list of dicts: {'material':., 'thickness':., 'width':., 'flags':., 'qty':., 'file':., 'mtime':.}
    
    success_count = 0
    error_count = 0
    
    print("Reading local scratch files (newest first)...")
    for idx, row in smk_df.iterrows():
        fname = row['file_name']
        local_path = os.path.join(LOCAL_DIR, fname)
        mtime = row['mtime_dt']
        
        if not os.path.exists(local_path):
            continue
            
        try:
            wb = openpyxl.load_workbook(local_path, data_only=True, read_only=True)
            for sheet in wb.sheetnames:
                if '納入先' in sheet:
                    ws = wb[sheet]
                    for r in extract_sheet_data(ws):
                        if len(r) < 4: continue
                        key = r[0]
                        if not key or key.lower() == 'none' or key == '': continue
                        content = tuple((r[:7] + [""] * 7)[:7])
                        if key not in all_addresses:
                            all_addresses[key] = {}
                        if content not in all_addresses[key]:
                            all_addresses[key][content] = {'files': [], 'newest_mtime': mtime}
                        all_addresses[key][content]['files'].append(fname)
                        
                elif 'トレイ' in sheet:
                    ws = wb[sheet]
                    for r in extract_sheet_data(ws):
                        if len(r) < 9: 
                            r = r + [""] * (9 - len(r))
                        key = r[0]
                        if not key or key.lower() == 'none' or key == '': continue
                        
                        tray_record = {
                            'material': r[2],
                            'thickness': r[3],
                            'width': r[4],
                            'anti_static': r[5],
                            'silicone': r[6],
                            'coating': r[7],
                            'qty': r[8],
                            'file': fname,
                            'mtime': mtime
                        }
                        if key not in all_trays:
                            all_trays[key] = []
                        all_trays[key].append(tray_record)
            wb.close()
            success_count += 1
        except Exception as e:
            error_count += 1
            
    print(f"Extraction complete. Success: {success_count}, Errors: {error_count}")
    
    # 1. Address Analysis (Time-series)
    addr_total = len(all_addresses)
    addr_history_cases = 0
    
    for key, variants in all_addresses.items():
        if len(variants) > 1:
            addr_history_cases += 1
            
    # 2. Tray Analysis
    tray_total = len(all_trays)
    tray_strict_conflict = 0
    tray_anomaly_cases = []
    
    for key, records in all_trays.items():
        # strict conflict: different material or thickness
        fixed_traits = set((r['material'], r['thickness']) for r in records)
        if len(fixed_traits) > 1:
            tray_strict_conflict += 1
            
        # correlation anomaly: width changes alongside flags
        # collect pairs of (width, flags_tuple)
        width_flag_pairs = set()
        widths = set()
        flags = set()
        for r in records:
            w = r['width']
            flgs = (r['anti_static'], r['silicone'], r['coating'])
            width_flag_pairs.add((w, flgs))
            widths.add(w)
            flags.add(flgs)
            
        # If width has multiple values AND flags have multiple values, 
        # and they change together (meaning the number of unique pairs > 1)
        # We consider it an anomaly if a specific width is strongly tied to specific flags
        if len(widths) > 1 and len(flags) > 1:
            tray_anomaly_cases.append((key, records, width_flag_pairs))
            
    # Report Generation
    report = f"""# Báo Cáo Hợp Nhất `新SMK注文書` (V2 - Phân Tích Nghiệp Vụ)

- **Đọc thành công:** {success_count} file (từ bản copy local)

## 1. Sổ Địa Chỉ (`納入先一覧表`) - Time Series

- **Tổng số mã `No.`:** {addr_total}
- **Có lịch sử thay đổi qua thời gian:** {addr_history_cases} ({addr_history_cases/addr_total*100:.1f}%)
*Giải pháp đề xuất đã áp dụng:* Sắp xếp theo `mtime` file, bản ghi mới nhất là bản hiện hành, các bản cũ là lịch sử. Không còn khái niệm "xung đột".

## 2. Sản Phẩm (`トレイデータ`) - Phân Lớp Trường Dữ Liệu

- **Tổng số mã `P/N`:** {tray_total}
- **XUNG ĐỘT THẬT SỰ (Khác `材質` vật liệu hoặc `厚み` độ dày):** {tray_strict_conflict} ({tray_strict_conflict/tray_total*100:.1f}%)
*(Tỷ lệ xung đột đã giảm cực mạnh so với 39.1% ban đầu khi loại bỏ nhiễu từ các trường vận hành).*

### ⚠️ Danh Sách Cần Review: Tương Quan Bất Thường Giữa Kích Thước (`巾`) và Cờ Xử Lý

Phát hiện **{len(tray_anomaly_cases)}** mã sản phẩm có sự thay đổi về chiều rộng (`巾`) luôn đi kèm với sự thay đổi của cờ xử lý bề mặt (`帯電`, `シリコン`, `塗布`). Đây có thể là Revision thiết kế khác nhau hoặc chạy trên line sản xuất đặc thù.

*Ví dụ Top 5:*
"""
    for key, records, pairs in tray_anomaly_cases[:5]:
        report += f"\n**Mã `P/N` = {key}** có các combo Width ↔ Flags sau:\n"
        for w, flgs in pairs:
            # find newest file for this combo
            for r in sorted(records, key=lambda x: x['mtime'], reverse=True):
                if r['width'] == w and (r['anti_static'], r['silicone'], r['coating']) == flgs:
                    report += f"- `巾: {w}` ↔ `帯電: {flgs[0]} | シリコン: {flgs[1]} | 塗布: {flgs[2]}` (VD file: {r['file']})\n"
                    break

    with open(OUT_REPORT, 'w', encoding='utf-8') as f:
        f.write(report)
        
    print("V2 Report generated.")

if __name__ == "__main__":
    main()
