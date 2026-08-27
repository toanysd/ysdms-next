import os
import shutil
import pandas as pd
import openpyxl
from datetime import datetime
import json

INVENTORY_CSV = r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\server_inventory.csv"
LOCAL_DIR = r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\scratch_smk"
OUT_REPORT = r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\SMK_Merge_Report.md"

def extract_sheet_data(ws, min_cols=5):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue # header
        row_strs = [str(cell).strip() if cell is not None else "" for cell in row]
        # Ignore completely empty rows
        if any(row_strs):
            # pad or slice to a specific number of columns to make them uniform?
            # 納入先一覧表 typically has: No., 送り先, 住所, 依頼元, サブ, SĐT, FAX
            rows.append(row_strs)
    return rows

def process_address_book(records, file_name, all_addresses):
    # record: [No, Dest, Addr, Req, Sub, Tel, Fax, ...]
    for r in records:
        if len(r) < 4: continue
        key = r[0] # No.
        if not key or key.lower() == 'none' or key == '': continue
        
        # We take first 7 columns for comparison: No, Dest, Addr, Req, Sub, Tel, Fax
        content = r[:7]
        # pad with empty strings if too short
        while len(content) < 7:
            content.append("")
            
        content_tuple = tuple(content)
        
        if key not in all_addresses:
            all_addresses[key] = {}
            
        if content_tuple not in all_addresses[key]:
            all_addresses[key][content_tuple] = []
            
        all_addresses[key][content_tuple].append(file_name)

def process_tray_data(records, file_name, all_trays):
    # P/N is usually column 0.
    for r in records:
        if len(r) < 4: continue
        key = r[0]
        if not key or key.lower() == 'none' or key == '': continue
        
        # Let's take the first 10 columns for comparison
        content = r[:10]
        while len(content) < 10:
            content.append("")
            
        content_tuple = tuple(content)
        
        if key not in all_trays:
            all_trays[key] = {}
            
        if content_tuple not in all_trays[key]:
            all_trays[key][content_tuple] = []
            
        all_trays[key][content_tuple].append(file_name)

def main():
    if not os.path.exists(LOCAL_DIR):
        os.makedirs(LOCAL_DIR)
        
    df = pd.read_csv(INVENTORY_CSV)
    
    # Filter strictly for 新SMK注文書. Since parent_dir might be garbled in print,
    # in the dataframe it's literally "新SMK注文書".
    smk_df = df[df['parent_dir'] == '新SMK注文書']
    if smk_df.empty:
        # Fallback to contains
        smk_df = df[df['full_path'].str.contains(r'\\新SMK注文書\\', case=False, na=False)]
        
    print(f"Found {len(smk_df)} files in 新SMK注文書.")
    
    all_addresses = {} # key -> { content_tuple: [file1, file2] }
    all_trays = {}
    
    success_count = 0
    error_count = 0
    
    print("Copying and extracting...")
    for idx, row in smk_df.iterrows():
        orig_path = row['full_path']
        fname = row['file_name']
        local_path = os.path.join(LOCAL_DIR, fname)
        
        try:
            # Copy file
            if not os.path.exists(local_path):
                shutil.copy2(orig_path, local_path)
            
            # Read read-only
            wb = openpyxl.load_workbook(local_path, data_only=True, read_only=True)
            
            # Sheets might have slightly different names, e.g., trailing spaces
            for sheet in wb.sheetnames:
                if '納入先' in sheet:
                    ws = wb[sheet]
                    recs = extract_sheet_data(ws)
                    process_address_book(recs, fname, all_addresses)
                elif 'トレイ' in sheet:
                    ws = wb[sheet]
                    recs = extract_sheet_data(ws)
                    process_tray_data(recs, fname, all_trays)
                    
            wb.close()
            success_count += 1
            if success_count % 100 == 0:
                print(f"Processed {success_count} / {len(smk_df)} files...")
                
        except Exception as e:
            error_count += 1
            # print(f"Error on {fname}: {e}")
            
    print(f"Extraction complete. Success: {success_count}, Errors: {error_count}")
    
    # Analyze conflicts - Addresses
    addr_total_keys = len(all_addresses)
    addr_no_conflict = 0
    addr_conflict = 0
    
    addr_conflict_examples = []
    for key, variants in all_addresses.items():
        if len(variants) == 1:
            addr_no_conflict += 1
        else:
            addr_conflict += 1
            if len(addr_conflict_examples) < 5:
                addr_conflict_examples.append((key, variants))
                
    # Analyze conflicts - Trays
    tray_total_keys = len(all_trays)
    tray_no_conflict = 0
    tray_conflict = 0
    
    tray_conflict_examples = []
    for key, variants in all_trays.items():
        if len(variants) == 1:
            tray_no_conflict += 1
        else:
            tray_conflict += 1
            if len(tray_conflict_examples) < 5:
                tray_conflict_examples.append((key, variants))
                
    # Generate Markdown Report
    report = f"""# Báo Cáo Hợp Nhất Thư Mục `新SMK注文書`

- **Tổng số file:** {len(smk_df)}
- **Đọc thành công:** {success_count}
- **Lỗi đọc file (hỏng/định dạng cũ):** {error_count}

## 1. Sheet `納入先一覧表` (Sổ Địa Chỉ)

- **Tổng số mã `No.` (Khách hàng/Điểm giao) tìm thấy:** {addr_total_keys}
- **Khớp hoàn toàn 100% nội dung (Không có xung đột):** {addr_no_conflict} ({addr_no_conflict/addr_total_keys*100:.1f}%)
- **CÓ XUNG ĐỘT (Cùng `No.` nhưng nội dung khác nhau):** {addr_conflict} ({addr_conflict/addr_total_keys*100:.1f}%)

### ⚠️ Ví dụ Xung Đột Điển Hình (Top 5):
"""
    for key, variants in addr_conflict_examples:
        report += f"\n**Mã `No.` = {key}** có {len(variants)} phiên bản khác nhau:\n"
        for i, (content, files) in enumerate(variants.items()):
            content_str = " | ".join(content)
            report += f"- *Phiên bản {i+1}* (xuất hiện trong {len(files)} file, ví dụ: `{files[0]}`):\n  `{content_str}`\n"
            
    report += f"\n## 2. Sheet `トレイデータ` (Sản Phẩm/Khay)\n\n"
    report += f"- **Tổng số mã `P/N` tìm thấy:** {tray_total_keys}\n"
    
    if tray_total_keys > 0:
        report += f"- **Khớp hoàn toàn 100% nội dung:** {tray_no_conflict} ({tray_no_conflict/tray_total_keys*100:.1f}%)\n"
        report += f"- **CÓ XUNG ĐỘT (Cùng `P/N` nhưng specs khác nhau):** {tray_conflict} ({tray_conflict/tray_total_keys*100:.1f}%)\n"
    
        report += f"\n### ⚠️ Ví dụ Xung Đột Điển Hình (Top 5):\n"
        for key, variants in tray_conflict_examples:
            report += f"\n**Mã `P/N` = {key}** có {len(variants)} phiên bản khác nhau:\n"
            for i, (content, files) in enumerate(variants.items()):
                content_str = " | ".join(content)
                report += f"- *Phiên bản {i+1}* (xuất hiện trong {len(files)} file, ví dụ: `{files[0]}`):\n  `{content_str}`\n"
                
    with open(OUT_REPORT, 'w', encoding='utf-8') as f:
        f.write(report)
        
    print(f"Report saved to {OUT_REPORT}")

if __name__ == "__main__":
    main()
