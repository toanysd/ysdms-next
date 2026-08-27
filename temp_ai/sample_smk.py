import os
import shutil
import glob
import pandas as pd
import openpyxl
from datetime import datetime

SERVER_PATH = r"\\SERVER\ysd-folder"
TARGET_FOLDER = "新SMK注文書"
LOCAL_DIR = r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\scratch_readonly"

def print_stat(path, label):
    stat = os.stat(path)
    mtime = datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
    ctime = datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
    print(f"[{label}] {os.path.basename(path)}")
    print(f"   Size: {stat.st_size} bytes")
    print(f"   MTime (Modified): {mtime}")
    print(f"   CTime (Created): {ctime}")
    print("-" * 50)

def main():
    if not os.path.exists(LOCAL_DIR):
        os.makedirs(LOCAL_DIR)
        
    df = pd.read_csv(r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\server_inventory.csv")
    
    # Filter for 新SMK注文書
    # We must match exactly the folder name in the dataframe
    # Because of shift-jis, let's just find the folder ending in 'SMK注文書'
    smk_df = df[df['parent_dir'].str.contains('SMK') & df['parent_dir'].str.contains('注')].copy()
    if smk_df.empty:
        # Fallback to wildcard search
        smk_df = df[df['full_path'].str.contains(r'SMK.*\.xls', case=False, na=False)].copy()
        
    # Find the specific target folder
    target_df = df[df['parent_dir'].str.contains('新SMK注', na=False)].copy()
    if target_df.empty:
        # manual match
        target_df = df[df['full_path'].str.contains('新SMK注', case=False, na=False)].copy()
        
    if target_df.empty:
        print("Target folder '新SMK注文書' not found in inventory. Using closest match.")
        target_df = smk_df
        
    # Get 3-5 newest files
    target_df['mtime_dt'] = pd.to_datetime(target_df['mtime'])
    target_df = target_df.sort_values('mtime_dt', ascending=False)
    
    top_5 = target_df.head(5)
    
    print("=== COPY AND VERIFY MTIME ===\n")
    local_files = []
    
    for _, row in top_5.iterrows():
        orig_path = row['full_path']
        fname = row['file_name']
        local_path = os.path.join(LOCAL_DIR, fname)
        
        # Copy preserving metadata
        shutil.copy2(orig_path, local_path)
        local_files.append(local_path)
        
        # Verify
        print_stat(orig_path, "ORIGINAL (SERVER)")
        print_stat(local_path, "COPIED (LOCAL)")
        
    print("\n=== READ EXCEL SHEETS (First 10 rows) ===\n")
    
    for lfile in local_files:
        print(f"\n>>> File: {os.path.basename(lfile)}")
        try:
            wb = openpyxl.load_workbook(lfile, data_only=True, read_only=True)
            print(f"Sheets: {wb.sheetnames}")
            
            for sheet in wb.sheetnames:
                print(f"\n--- Sheet: {sheet} ---")
                ws = wb[sheet]
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 10: break
                    row_strs = [str(cell) for cell in row]
                    # Filter out purely None rows to save space
                    if any(c != 'None' for c in row_strs):
                        print(f"Row {i+1}: " + " | ".join(row_strs))
                        
            wb.close()
        except Exception as e:
            print(f"Error reading {lfile}: {e}")

if __name__ == "__main__":
    main()
