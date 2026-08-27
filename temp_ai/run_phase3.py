import os
import shutil
import pandas as pd
import openpyxl
from datetime import datetime
import json
import hashlib
import traceback
import sys
import argparse

INVENTORY_CSV = r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\server_inventory.csv"
SCRATCH_BASE = r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\scratch_phase3"
REPORTS_DIR = r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\reports_phase3"
PARSED_DIR = r"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\parsed_phase3"

for d in [SCRATCH_BASE, REPORTS_DIR, PARSED_DIR]:
    if not os.path.exists(d):
        os.makedirs(d)

def safe_hash(s):
    return hashlib.md5(s.encode('utf-8', errors='ignore')).hexdigest()[:12]

def extract_sheet_data(ws):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        row_strs = [str(cell).strip() if cell is not None else "" for cell in row]
        if any(row_strs):
            rows.append(row_strs)
    return rows

def process_folder(parent_dir, files_df):
    folder_hash = safe_hash(parent_dir)
    local_dir = os.path.join(SCRATCH_BASE, folder_hash)
    if not os.path.exists(local_dir):
        os.makedirs(local_dir)
        
    all_addresses = {} 
    all_trays = {} 
    success = 0
    errors = 0
    
    files_df = files_df.sort_values('mtime_dt', ascending=False)
    
    for idx, row in files_df.iterrows():
        orig_path = row['full_path']
        fname = row['file_name']
        local_path = os.path.join(local_dir, fname)
        mtime = row['mtime_dt']
        
        try:
            shutil.copy2(orig_path, local_path)
            wb = openpyxl.load_workbook(local_path, data_only=True, read_only=True)
            for sheet in wb.sheetnames:
                if '納入先' in sheet:
                    ws = wb[sheet]
                    for r in extract_sheet_data(ws):
                        if len(r) < 4: continue
                        key = r[0]
                        if not key or key.lower() == 'none' or key == '': continue
                        content = tuple((r[:7] + [""] * 7)[:7])
                        if key not in all_addresses:
                            all_addresses[key] = {}
                        if content not in all_addresses[key]:
                            all_addresses[key][content] = {'files': [], 'newest_mtime': mtime.isoformat()}
                        all_addresses[key][content]['files'].append(fname)
                        
                elif 'トレイ' in sheet:
                    ws = wb[sheet]
                    for r in extract_sheet_data(ws):
                        if len(r) < 9: r = r + [""] * (9 - len(r))
                        key = r[0]
                        if not key or key.lower() == 'none' or key == '': continue
                        tray_record = {
                            'material': r[2], 'thickness': r[3], 'width': r[4],
                            'anti_static': r[5], 'silicone': r[6], 'coating': r[7],
                            'qty': r[8], 'file': fname, 'mtime': mtime.isoformat()
                        }
                        if key not in all_trays:
                            all_trays[key] = []
                        all_trays[key].append(tray_record)
            wb.close()
            success += 1
        except Exception as e:
            errors += 1
            
    try:
        shutil.rmtree(local_dir)
    except:
        pass
        
    addr_total = len(all_addresses)
    addr_history = sum(1 for v in all_addresses.values() if len(v) > 1)
    
    tray_total = len(all_trays)
    tray_strict_conflict = 0
    tray_anomaly_cases = []
    
    for key, records in all_trays.items():
        fixed_traits = set((r['material'], r['thickness']) for r in records)
        if len(fixed_traits) > 1:
            tray_strict_conflict += 1
            
        widths = set()
        flags = set()
        for r in records:
            w = r['width']
            flg = (r['anti_static'], r['silicone'], r['coating'])
            widths.add(w)
            flags.add(flg)
            
        if len(widths) > 1 and len(flags) > 1:
            tray_anomaly_cases.append(key)
            
    out_data = {
        'parent_dir': parent_dir,
        'success': success,
        'errors': errors,
        'addr_total': addr_total,
        'addr_history': addr_history,
        'tray_total': tray_total,
        'tray_strict_conflict': tray_strict_conflict,
        'tray_anomalies': len(tray_anomaly_cases)
    }
    with open(os.path.join(PARSED_DIR, f"{folder_hash}.json"), 'w', encoding='utf-8') as f:
        json.dump(out_data, f, ensure_ascii=False)
        
    report = f"# Báo Cáo Thư Mục: `{parent_dir}`\n"
    report += f"- Đọc thành công: {success}, Lỗi: {errors}\n"
    report += f"- Sổ địa chỉ: {addr_total} mã (Lịch sử: {addr_history})\n"
    report += f"- Sản phẩm: {tray_total} mã (Xung đột cố định: {tray_strict_conflict}, Anomaly Width/Flags: {len(tray_anomaly_cases)})\n"
    
    with open(os.path.join(REPORTS_DIR, f"{folder_hash}.md"), 'w', encoding='utf-8') as f:
        f.write(report)
        
    return out_data

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--limit', type=int, default=10, help='Max folders to process')
    parser.add_argument('--offset', type=int, default=0, help='Folder index offset')
    args = parser.parse_args()
    
    df = pd.read_csv(INVENTORY_CSV)
    df = df[~df['file_name'].str.lower().str.endswith('.xls')].copy()
    df['mtime_dt'] = pd.to_datetime(df['mtime'])
    
    # Sort folders by max mtime (newest activity first)
    folder_activity = df.groupby('parent_dir')['mtime_dt'].max().reset_index()
    folder_activity = folder_activity.sort_values('mtime_dt', ascending=False)
    
    active_folders = folder_activity['parent_dir'].tolist()
    
    start_idx = args.offset
    end_idx = min(start_idx + args.limit, len(active_folders))
    target_folders = active_folders[start_idx:end_idx]
    
    print(f"Starting Phase 3 Pipeline for {len(target_folders)} folders (Offset {start_idx} to {end_idx})...")
    
    grand_stats = {
        'folders_processed': 0,
        'folders_failed': 0,
        'files_success': 0,
        'files_error': 0,
        'addr_total': 0,
        'addr_history': 0,
        'tray_total': 0,
        'tray_strict_conflict': 0,
        'tray_anomalies': 0
    }
    
    batch_report = f"# Báo Cáo Lô Phase 3 (Offset {start_idx} - {end_idx})\n\n| Thư mục | File OK | Lỗi | KH | KH Lịch sử | Sản phẩm | SP Xung đột | SP Anomaly |\n|---|---|---|---|---|---|---|---|\n"
    
    for i, pdir in enumerate(target_folders):
        print(f"[{i+1}/{len(target_folders)}] Processing {pdir}...")
        sys.stdout.flush()
        
        folder_df = df[df['parent_dir'] == pdir].copy()
        try:
            stats = process_folder(pdir, folder_df)
            grand_stats['folders_processed'] += 1
            grand_stats['files_success'] += stats['success']
            grand_stats['files_error'] += stats['errors']
            grand_stats['addr_total'] += stats['addr_total']
            grand_stats['addr_history'] += stats['addr_history']
            grand_stats['tray_total'] += stats['tray_total']
            grand_stats['tray_strict_conflict'] += stats['tray_strict_conflict']
            grand_stats['tray_anomalies'] += stats['tray_anomalies']
            
            batch_report += f"| `{pdir}` | {stats['success']} | {stats['errors']} | {stats['addr_total']} | {stats['addr_history']} | {stats['tray_total']} | {stats['tray_strict_conflict']} | {stats['tray_anomalies']} |\n"
        except Exception as e:
            print(f"ERROR processing folder {pdir}: {e}")
            grand_stats['folders_failed'] += 1
            batch_report += f"| `{pdir}` | LỖI | LỖI | - | - | - | - | - |\n"
            
    batch_report += f"\n## TỔNG KẾT LÔ\n- Thư mục xong: {grand_stats['folders_processed']}\n- Thư mục lỗi: {grand_stats['folders_failed']}\n- File xong: {grand_stats['files_success']}\n- File lỗi: {grand_stats['files_error']}\n"
    
    with open(rf"d:\AntiGravity_Workspace\apps\ysdms-nextgen\temp_ai\Phase3_Batch_{start_idx}_to_{end_idx}.md", 'w', encoding='utf-8') as f:
        f.write(batch_report)
        
    print(f"BATCH COMPLETE! Saved report to Phase3_Batch_{start_idx}_to_{end_idx}.md")

if __name__ == "__main__":
    main()
