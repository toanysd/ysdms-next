import sys, os

def dump_xlsx(f, out):
    import openpyxl
    with open(out, 'w', encoding='utf-8') as fh:
        wb = openpyxl.load_workbook(f, data_only=True)
        fh.write(f'File: {os.path.basename(f)}\n')
        for name in wb.sheetnames:
            ws = wb[name]
            fh.write(f'Sheet: {name} (up to 25 rows)\n')
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True)):
                vals = []
                for c_idx, v in enumerate(row):
                    if v is not None and str(v).strip():
                        vals.append(f'[{c_idx}]{v}')
                if vals:
                    fh.write(f'  R{r_idx+1}: {" | ".join(vals)}\n')

if __name__ == '__main__':
    dump_xlsx(sys.argv[1], sys.argv[2])
