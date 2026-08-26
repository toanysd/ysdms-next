import xlrd, os, glob, sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def parse_xls(filepath):
    try:
        print(f"\n{'='*60}")
        print(f"FILE: {os.path.basename(filepath)}")
        if filepath.endswith('.xls'):
            wb = xlrd.open_workbook(filepath, encoding_override='shift-jis')
            print(f"Sheets: {wb.sheet_names()}")
            for name in wb.sheet_names()[:5]:
                ws = wb.sheet_by_name(name)
                print(f"\n  Sheet: {name} ({ws.nrows} rows x {ws.ncols} cols)")
                for r in range(min(20, ws.nrows)):
                    vals = []
                    for c in range(min(20, ws.ncols)):
                        v = ws.cell_value(r, c)
                        if v and str(v).strip():
                            vals.append(f"[{c}]{v}".replace('\n', ' '))
                    if vals:
                        print(f"    Row {r+1}: {' | '.join(vals)}")
        elif filepath.endswith('.xlsx'):
            wb = openpyxl.load_workbook(filepath, data_only=True)
            print(f"Sheets: {wb.sheetnames}")
            for name in wb.sheetnames[:5]:
                ws = wb[name]
                print(f"\n  Sheet: {name} ({ws.max_row} rows x {ws.max_column} cols)")
                for r in range(1, min(21, ws.max_row + 1)):
                    vals = []
                    for c in range(1, min(21, ws.max_column + 1)):
                        v = ws.cell(row=r, column=c).value
                        if v and str(v).strip():
                            vals.append(f"[{c-1}]{v}".replace('\n', ' '))
                    if vals:
                        print(f"    Row {r}: {' | '.join(vals)}")
    except Exception as e:
        print(f"ERROR {filepath}: {e}")

# Parse all xls files in Form lien quan
form_dir = r'D:\AntiGravity_Workspace\apps\ysdms-nextgen\source_data\Form lien quan'
for f in glob.glob(os.path.join(form_dir, '*.xls*')):
    if any(keyword in f for keyword in ['日報', '記録', 'report', '指示']):
        parse_xls(f)

# Parse ISO format files
iso_dir = r'D:\AntiGravity_Workspace\apps\ysdms-nextgen\source_data\ISO(2026見直し済み）\文書類　共通\フォーマット'
for f in glob.glob(os.path.join(iso_dir, '*.xls*')):
    if any(keyword in f for keyword in ['日報', '記録', 'report', '指示', '工程表', '成形条件', '不適合', '出荷']):
        parse_xls(f)
